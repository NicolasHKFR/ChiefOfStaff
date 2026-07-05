import csv
import io

import openpyxl
from openpyxl.styles import Font


def export_csv(columns, rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def export_xlsx(columns, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
