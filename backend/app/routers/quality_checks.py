import json
import os
import uuid

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import settings
from app.database import get_db
from app.models.quality_check import QualityCheck, QCFile
from app.schemas.common import QualityCheckCreate, QualityCheckDetail, QualityCheckOut, QCFileOut

router = APIRouter(prefix="/quality-checks", tags=["Quality Checks"])


def _parse_excel(filepath: str) -> list[dict]:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    result = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]
        data_rows = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            data_rows.append({headers[i]: cell for i, cell in enumerate(row)})
        result.append({
            "sheet_name": sheet_name,
            "headers": headers,
            "rows": data_rows,
        })
    wb.close()
    return result


@router.get("", response_model=list[QualityCheckOut])
async def list_quality_checks(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(QualityCheck).order_by(QualityCheck.created_at.desc()))
    return result.scalars().all()


@router.post("", response_model=QualityCheckOut, status_code=201)
async def create_quality_check(data: QualityCheckCreate, db: AsyncSession = Depends(get_db)):
    qc = QualityCheck(name=data.name, description=data.description)
    db.add(qc)
    await db.commit()
    await db.refresh(qc)
    return qc


@router.get("/{qc_id}", response_model=QualityCheckDetail)
async def get_quality_check(qc_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(QualityCheck)
        .where(QualityCheck.id == qc_id)
        .options(selectinload(QualityCheck.files), selectinload(QualityCheck.checks))
    )
    qc = result.scalar_one_or_none()
    if not qc:
        raise HTTPException(404, "Quality check not found")
    return qc


@router.delete("/{qc_id}", status_code=204)
async def delete_quality_check(qc_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(QualityCheck)
        .where(QualityCheck.id == qc_id)
        .options(selectinload(QualityCheck.files))
    )
    qc = result.scalar_one_or_none()
    if not qc:
        raise HTTPException(404, "Quality check not found")
    for f in qc.files:
        if os.path.exists(f.stored_path):
            os.remove(f.stored_path)
    await db.delete(qc)
    await db.commit()


@router.post("/{qc_id}/upload", response_model=QCFileOut, status_code=201)
async def upload_file(qc_id: int, file: UploadFile, db: AsyncSession = Depends(get_db)):
    qc = await db.get(QualityCheck, qc_id)
    if not qc:
        raise HTTPException(404, "Quality check not found")

    ext = os.path.splitext(file.filename or "file")[1] or ".xlsx"
    filename = f"{uuid.uuid4().hex}{ext}"
    qc_upload_dir = os.path.join(settings.upload_dir, "qc")
    os.makedirs(qc_upload_dir, exist_ok=True)
    filepath = os.path.join(qc_upload_dir, filename)

    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    try:
        parsed = _parse_excel(filepath)
        data_json = json.loads(json.dumps(parsed, default=str))
        row_count = sum(len(sheet["rows"]) for sheet in parsed)
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        raise HTTPException(400, f"Failed to parse Excel file: {e}")

    qc_file = QCFile(
        quality_check_id=qc_id,
        original_filename=file.filename or "unknown.xlsx",
        stored_path=filepath,
        data_json=data_json,
        row_count=row_count,
    )
    qc.status = "in_progress"
    db.add(qc_file)
    await db.commit()
    await db.refresh(qc_file)
    return qc_file


@router.get("/{qc_id}/files/{file_id}/download")
async def download_file(qc_id: int, file_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(QCFile).where(QCFile.id == file_id, QCFile.quality_check_id == qc_id)
    )
    qc_file = result.scalar_one_or_none()
    if not qc_file:
        raise HTTPException(404, "File not found")
    if not os.path.exists(qc_file.stored_path):
        raise HTTPException(404, "File not found on disk")
    from fastapi.responses import FileResponse
    return FileResponse(qc_file.stored_path, filename=qc_file.original_filename)
